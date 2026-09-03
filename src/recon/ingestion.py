"""Carregamento de CSV/XLSX/XLS/XLSB com detecção de encoding e separador."""
import bz2
import csv
import gzip
import lzma
import os
import zipfile
from collections.abc import Sequence
from typing import Any, Literal

import numpy as np
import pandas as pd
from charset_normalizer import from_bytes
from loguru import logger

from . import layout as layout_mod

ExcelEngine = Literal["xlrd", "openpyxl", "pyxlsb"]

_ENGINES_EXCEL: dict[str, ExcelEngine] = {".xlsx": "openpyxl", ".xls": "xlrd", ".xlsb": "pyxlsb"}
_SEPARADORES_CANDIDATOS: Sequence[str] = (",", ";", "\t", "|")
_LINHAS_AMOSTRA_SNIFF = 50
_MAX_MEMBROS_ZIP = 1
_MAX_TAMANHO_DESCOMPACTADO_ZIP = 10 * 1024**3
_MAX_RAZAO_COMPRESSAO_ZIP = 250

# Texto separado por delimitador, qualquer que seja o nome da extensão. O
# detector de separador já resolve tabulação e ponto e vírgula — recusar `.tsv`
# e `.txt` por causa da extensão era barrar arquivo que a ferramenta já sabia
# ler. As compactadas o pandas abre sozinho, pelo sufixo.
_EXTENSOES_TEXTO = frozenset({".csv", ".tsv", ".txt"})
_COMPACTADAS = (".gz", ".bz2", ".zip", ".xz", ".zst")
# Aceitas quando o usuário aponta o arquivo.
EXTENSOES_SUPORTADAS: tuple[str, ...] = (
    ".csv", ".tsv", ".txt", ".xlsx", ".xls", ".xlsb", ".parquet",
    ".csv.gz", ".tsv.gz", ".txt.gz", ".csv.zip",
)
# Procuradas ao varrer uma pasta. `.txt` fica de fora de propósito: pasta de
# trabalho tem `leiame.txt`, `notas.txt` e log de sistema, e varrer tudo isso
# como se fosse tabela enche o relatório de lixo. Apontado na mão, continua
# valendo — quem escolheu o arquivo sabe o que ele é.
EXTENSOES_DESCOBERTAS: tuple[str, ...] = tuple(
    e for e in EXTENSOES_SUPORTADAS if not e.startswith(".txt")
)


def _partes_da_extensao(caminho: str) -> tuple[str, str]:
    """(extensão lógica, extensão de compactação) de um caminho.

    `vendas.csv.gz` devolve `('.csv', '.gz')`; `vendas.xlsx`, `('.xlsx', '')`.
    """
    nome = os.path.basename(caminho).lower()
    compactacao = next((c for c in _COMPACTADAS if nome.endswith(c)), "")
    if compactacao:
        nome = nome[: -len(compactacao)]
    return os.path.splitext(nome)[1], compactacao


def formato_de(caminho: str) -> str:
    """`texto`, `excel`, `parquet` ou `desconhecido`."""
    extensao, _ = _partes_da_extensao(caminho)
    if extensao in _EXTENSOES_TEXTO:
        return "texto"
    if extensao in _ENGINES_EXCEL:
        return "excel"
    if extensao == ".parquet":
        return "parquet"
    return "desconhecido"


class IngestionError(Exception):
    """Base para todas as falhas tipadas de ingestão."""


class FileFormatError(IngestionError):
    """Extensão de arquivo não suportada pelo profiler."""


class EncodingDetectionError(IngestionError):
    """Falha ao detectar o encoding de um arquivo CSV."""


def _validar_zip(caminho: str) -> zipfile.ZipInfo:
    """Aceita um ZIP simples e limita expansão hostil antes do pandas abrir.

    O Recon espera um único CSV compactado. Aceitar vários membros deixaria a
    escolha ambígua; aceitar razão/tamanho arbitrários abre espaço para ZIP
    bombs que consomem disco, CPU ou memória ao descompactar.
    """
    try:
        with zipfile.ZipFile(caminho) as arquivo:
            membros = [m for m in arquivo.infolist() if not m.is_dir()]
    except zipfile.BadZipFile as e:
        raise FileFormatError(f"Arquivo ZIP inválido: '{caminho}'.") from e
    if len(membros) != _MAX_MEMBROS_ZIP:
        raise FileFormatError(
            "ZIPs aceitos pelo Recon devem conter exatamente um arquivo de dados; "
            f"'{caminho}' contém {len(membros)} membro(s)."
        )
    membro = membros[0]
    if not membro.filename.lower().endswith(tuple(_EXTENSOES_TEXTO)):
        raise FileFormatError(
            f"O ZIP deve conter CSV/TSV/TXT; encontrou '{membro.filename}'."
        )
    if membro.file_size > _MAX_TAMANHO_DESCOMPACTADO_ZIP:
        raise FileFormatError(
            f"O conteúdo de '{caminho}' descompactaria para mais de 10 GB; "
            "recusei a leitura por segurança."
        )
    razao = membro.file_size / max(membro.compress_size, 1)
    if razao > _MAX_RAZAO_COMPRESSAO_ZIP:
        raise FileFormatError(
            f"O ZIP '{caminho}' tem razão de compressão incomum ({razao:.0f}×) e foi recusado "
            "por segurança."
        )
    return membro


# Bytes lidos para adivinhar o encoding. Antes o detector lia o arquivo
# inteiro: num CSV de 1 GB isso é um passo caro antes de qualquer análise, e a
# resposta não melhora depois dos primeiros parágrafos. Ler por amostra também
# é o que permite detectar dentro de um arquivo compactado.
_BYTES_AMOSTRA_ENCODING = 256_000


def _amostra_bytes(caminho: str, compactacao: str = "", limite: int = _BYTES_AMOSTRA_ENCODING) -> bytes:
    """Primeiros bytes já descompactados do arquivo."""
    if compactacao == ".gz":
        with gzip.open(caminho, "rb") as f:
            return f.read(limite)
    if compactacao == ".bz2":
        with bz2.open(caminho, "rb") as f:
            return f.read(limite)
    if compactacao in (".xz", ".zst"):
        with lzma.open(caminho, "rb") as f:
            return f.read(limite)
    if compactacao == ".zip":
        with zipfile.ZipFile(caminho) as z:
            membro = _validar_zip(caminho)
            with z.open(membro) as f:
                return f.read(limite)
    with open(caminho, "rb") as f:
        return f.read(limite)


def detectar_encoding(caminho: str, compactacao: str = "") -> str:
    try:
        amostra = _amostra_bytes(caminho, compactacao)
        resultado = from_bytes(amostra).best() if amostra else None
        if resultado is None:
            raise EncodingDetectionError(f"Não foi possível detectar encoding de '{caminho}'")
        encoding = resultado.encoding or "utf-8"
        logger.info(f"Encoding detectado: '{encoding}'")
        return encoding
    except EncodingDetectionError:
        raise
    except Exception as e:
        logger.warning(f"Falha na detecção de encoding: {e}. Usando utf-8.")
        return "utf-8"


def detectar_separador(caminho: str, encoding: str, compactacao: str = "") -> str:
    """Escolhe o separador pela consistência do número de campos por linha.

    A heurística anterior aceitava o primeiro separador que produzisse mais de
    uma coluna, e caía num sniffer genérico quando nenhum produzia. Isso
    corrompia CSV de coluna única em silêncio: `nome\\nAna\\nBruno` virava duas
    colunas `['n', 'me']`, com a letra "o" eleita separador.

    Aqui um separador só vence se todas as linhas da amostra se dividirem no
    mesmo número de campos, e uma única coluna é um resultado legítimo, não um
    sinal de falha.
    """
    try:
        texto = _amostra_bytes(caminho, compactacao).decode(encoding, errors="replace")
    except (OSError, LookupError, zipfile.BadZipFile) as e:
        raise FileFormatError(f"Falha ao ler '{caminho}' para detectar o separador: {e}") from e
    amostra = [
        linha + "\n"
        for linha in texto.splitlines()[:_LINHAS_AMOSTRA_SNIFF]
        if linha.strip()
    ]

    if not amostra:
        return ","

    melhor_sep = ","
    melhor_chave = (0.0, 0)
    for sep in _SEPARADORES_CANDIDATOS:
        try:
            linhas = [linha for linha in csv.reader(amostra, delimiter=sep) if linha]
        except csv.Error:
            continue
        if not linhas:
            continue
        contagens = [len(linha) for linha in linhas]
        # A largura típica é a moda, não a da primeira linha: num arquivo com
        # título ("RELATÓRIO DE PESSOAL") a primeira linha tem um campo só com
        # qualquer separador, e o candidato certo era descartado logo de saída —
        # o arquivo inteiro acabava lido como coluna única.
        n_campos = max(set(contagens), key=contagens.count)
        if n_campos < 2:
            continue
        consistencia = contagens.count(n_campos) / len(contagens)
        chave = (round(consistencia, 3), n_campos)
        if chave > melhor_chave:
            melhor_chave, melhor_sep = chave, sep

    if melhor_chave[1] < 2:
        logger.info("Nenhum separador produz mais de uma coluna — tratando como CSV de coluna única.")
        return ","
    return melhor_sep


def _ler_csv(caminho: str, encoding: str, sep: str) -> pd.DataFrame:
    """Lê o CSV preferindo o engine pyarrow (multi-thread, ~10× mais rápido
    que o engine C nesta carga), com queda para o engine C quando o pyarrow
    recusa o arquivo — ele é bem mais estrito com CSV malformado, que é o
    caso comum nos arquivos analisados aqui."""
    try:
        return pd.read_csv(caminho, encoding=encoding, sep=sep, engine="pyarrow")
    except Exception as e:
        logger.debug(f"Engine pyarrow recusou o arquivo ({e}); usando o engine C.")
        # `encoding_errors="replace"` porque a amostra que decide o encoding
        # olha só o início do arquivo — um export de sistema legado pode ter
        # uma dúzia de bytes corrompidos lá pelo meio (o mesmo byte virando
        # letras diferentes em palavras diferentes é sinal de que a fonte já
        # estava quebrada antes de chegar aqui). Travar a análise inteira de
        # um arquivo de 100+ MB por uma dúzia de bytes é pior que substituir
        # por "�" e avisar — `_avisar_se_encoding_teve_substituicao` cuida do
        # aviso.
        return pd.read_csv(
            caminho, encoding=encoding, sep=sep, low_memory=False, encoding_errors="replace"
        )


_LINHAS_INSPECAO_LAYOUT = 40

# Acima deste tamanho, o CSV é lido em blocos e amostrado durante a leitura.
# O caminho normal carrega o arquivo inteiro e só depois amostra, o que custa
# 5-6× o tamanho do arquivo em RAM — um CSV de 2 GB simplesmente não abre numa
# máquina de 8 GB, e é justamente onde a amostra seria mais necessária.
TAMANHO_LEITURA_EM_BLOCOS = 300 * 1024 * 1024
_LINHAS_POR_BLOCO = 200_000
_FRACAO_RAM_SEGURA = 0.70
_FATOR_MEMORIA_TEXTO = 6
_FATOR_MEMORIA_EXCEL = 8


def _memoria_sistema_bytes() -> tuple[int | None, int | None]:
    """Retorna (RAM total, RAM disponível), sem criar dependência externa."""
    try:
        total = os.sysconf("SC_PAGE_SIZE") * os.sysconf("SC_PHYS_PAGES")
        disponivel = total
        with open("/proc/meminfo", encoding="utf-8") as arquivo:
            campos = {
                linha.split(":", 1)[0]: int(linha.split()[1]) * 1024
                for linha in arquivo if ":" in linha
            }
        disponivel = campos.get("MemAvailable", disponivel)
        return total, disponivel
    except (AttributeError, OSError, ValueError):
        pass
    # A maior parte dos usuários de planilhas está em Windows; `sysconf` e
    # /proc não existem lá. Usar a API nativa mantém a mesma margem de 30%.
    if os.name == "nt":
        try:
            import ctypes

            class _StatusMemoria(ctypes.Structure):
                _fields_ = [
                    ("dwLength", ctypes.c_ulong), ("dwMemoryLoad", ctypes.c_ulong),
                    ("ullTotalPhys", ctypes.c_ulonglong), ("ullAvailPhys", ctypes.c_ulonglong),
                    ("ullTotalPageFile", ctypes.c_ulonglong), ("ullAvailPageFile", ctypes.c_ulonglong),
                    ("ullTotalVirtual", ctypes.c_ulonglong), ("ullAvailVirtual", ctypes.c_ulonglong),
                    ("ullAvailExtendedVirtual", ctypes.c_ulonglong),
                ]

            estado = _StatusMemoria()
            estado.dwLength = ctypes.sizeof(estado)
            windll: Any = ctypes.__dict__["windll"]
            if windll.kernel32.GlobalMemoryStatusEx(ctypes.byref(estado)):
                return int(estado.ullTotalPhys), int(estado.ullAvailPhys)
        except (AttributeError, OSError):
            pass
    # macOS não expõe `MemAvailable` como Linux. Sem uma fonte confiável de
    # memória livre, desabilitar a estimativa é mais honesto que inventar RAM.
    return None, None


def _amostragem_por_memoria(caminho: str, formato: str) -> str | None:
    """Explica quando uma leitura integral ultrapassaria o orçamento de RAM."""
    total, disponivel = _memoria_sistema_bytes()
    if total is None or disponivel is None:
        return None
    fator = _FATOR_MEMORIA_EXCEL if formato == "excel" else _FATOR_MEMORIA_TEXTO
    tamanho_base = os.path.getsize(caminho)
    # XLSX é um ZIP: olhar apenas seu tamanho comprimido pode subestimar muito
    # a memória necessária para XMLs, strings compartilhadas e DataFrame.
    if formato == "excel" and caminho.lower().endswith(".xlsx"):
        try:
            with zipfile.ZipFile(caminho) as livro:
                tamanho_base = max(tamanho_base, sum(info.file_size for info in livro.infolist()))
        except zipfile.BadZipFile:
            # A leitura normal vai produzir a mensagem apropriada para arquivo
            # corrompido; aqui basta voltar ao tamanho físico disponível.
            pass
    estimativa = tamanho_base * fator
    # A referência é a RAM disponível agora — não a RAM física total. Assim,
    # numa máquina de 16 GB já ocupada por outros programas, o Recon continua
    # deixando 30% do que realmente está livre para o sistema e processos de
    # fundo, em vez de assumir que os 16 GB inteiros estão à disposição.
    orcamento = int(disponivel * _FRACAO_RAM_SEGURA)
    if estimativa <= orcamento:
        return None
    return (
        f"Leitura integral estimada em {estimativa / 1024**3:.1f} GB, acima do orçamento "
        f"seguro de {orcamento / 1024**3:.1f} GB (70% da RAM disponível). "
        "O Recon usou uma amostra para evitar esgotar a memória."
    )


def _ler_csv_amostrado(
    caminho: str, encoding: str, sep: str, skiprows: int, limite: int
) -> tuple[pd.DataFrame, int]:
    """Lê o CSV em blocos e devolve (amostra, total de linhas do arquivo).

    A amostra é uniforme sobre todas as linhas, sem carregar o arquivo
    inteiro: cada linha recebe uma chave aleatória determinística e só as
    ``limite`` menores chaves continuam na memória. Isso evita o viés de
    analisar apenas o começo de exportações ordenadas por data.
    """
    total = 0
    pedacos: list[pd.DataFrame] = []
    chaves_amostra = np.empty(0, dtype=float)
    gerador = np.random.default_rng(42)
    leitor = pd.read_csv(
        caminho, encoding=encoding, sep=sep, skiprows=skiprows or None,
        chunksize=_LINHAS_POR_BLOCO, low_memory=False, encoding_errors="replace",
    )
    for bloco in leitor:
        total += len(bloco)
        chaves_bloco = gerador.random(len(bloco))
        candidatos = pedacos + [bloco]
        chaves_candidatas = np.concatenate((chaves_amostra, chaves_bloco))
        quantidade = min(limite, len(chaves_candidatas))
        indices = np.argpartition(chaves_candidatas, quantidade - 1)[:quantidade]
        combinado: pd.DataFrame = pd.concat(candidatos, ignore_index=True)
        pedacos = [combinado.iloc[indices].reset_index(drop=True)]
        chaves_amostra = chaves_candidatas[indices]

    df = pd.concat(pedacos, ignore_index=True) if pedacos else pd.DataFrame()
    logger.info(
        f"Leitura em blocos: {total:,} linhas no arquivo, {len(df):,} carregadas para análise."
    )
    return df, total


def _matriz_crua_csv(caminho: str, encoding: str, sep: str, compactacao: str = "") -> pd.DataFrame:
    """Primeiras linhas do CSV como matriz, sem interpretar cabeçalho.

    `read_csv(header=None)` fixa a largura pela primeira linha — num arquivo
    com título de uma célula, todas as linhas seguintes viram "linha ruim" e
    são descartadas, e a detecção de layout ficaria cega. Ler com `csv.reader`
    e preencher até a largura máxima preserva o formato real do arquivo.
    """
    try:
        texto = _amostra_bytes(caminho, compactacao).decode(encoding, errors="replace")
        cruas = texto.splitlines()[:_LINHAS_INSPECAO_LAYOUT]
        linhas = list(csv.reader(cruas, delimiter=sep))
    except (OSError, csv.Error, LookupError, zipfile.BadZipFile):
        return pd.DataFrame()
    if not linhas:
        return pd.DataFrame()
    largura = max(len(linha) for linha in linhas)
    normalizadas = [
        [(c.strip() or None) for c in linha] + [None] * (largura - len(linha))
        for linha in linhas
    ]
    return pd.DataFrame(normalizadas)


def _avisar_se_encoding_teve_substituicao(
    df: pd.DataFrame, avisos: list, encoding: str
) -> None:
    """Verifica se sobrou caractere de substituição ("�") depois da leitura
    tolerante e, se sim, deixa isso visível no relatório em vez de silencioso.

    `encoding_errors="replace"` evita a análise inteira travar por um byte
    corrompido, mas trocar sem avisar esconderia esse defeito de dado em vez
    de reportá-lo.
    """
    colunas_texto = df.select_dtypes(include=["object", "str"]).columns
    if colunas_texto.empty:
        return
    afetadas = [
        str(c) for c in colunas_texto
        if df[c].astype(str).str.contains("�", regex=False).any()
    ]
    if not afetadas:
        return
    avisos.append({
        "tipo": "encoding_substituido",
        "severidade": "🟡 MÉDIA",
        "mensagem": (
            f"Byte que não decodifica em '{encoding}' foi substituído por "
            f"\"�\" em {len(afetadas)} coluna(s): {', '.join(afetadas[:6])}"
            f"{'…' if len(afetadas) > 6 else ''}. Provável origem: bytes "
            "corrompidos no arquivo de origem, não erro de detecção — "
            "confira o valor original na fonte antes de usar essas colunas."
        ),
    })


def _anexar_layout(df: pd.DataFrame, lay: layout_mod.Layout) -> pd.DataFrame:
    """Guarda o diagnóstico de layout no próprio DataFrame.

    Via `attrs` para não mudar a assinatura de `carregar_arquivo` — o pipeline
    lê de lá na hora de montar o payload.
    """
    df.attrs["layout"] = lay
    return df


def _preparar_corpo(
    df: pd.DataFrame, avisos_iniciais: list
) -> tuple[pd.DataFrame, layout_mod.Layout]:
    df, lay = layout_mod.analisar_corpo(df)
    lay.avisos = list(avisos_iniciais) + lay.avisos
    return df, lay


def _carregar_csv_com_layout(
    caminho: str, detectar: bool, linha_cabecalho: int | None, limite_linhas: int | None = None
) -> pd.DataFrame:
    _, compactacao = _partes_da_extensao(caminho)
    encoding = detectar_encoding(caminho, compactacao)
    sep = detectar_separador(caminho, encoding, compactacao)

    inicio = linha_cabecalho or 0
    avisos: list = []
    if detectar and linha_cabecalho is None:
        inicio, avisos = layout_mod.detectar_linha_cabecalho(
            _matriz_crua_csv(caminho, encoding, sep, compactacao)
        )

    aviso_memoria = _amostragem_por_memoria(caminho, "texto") if limite_linhas is not None else None
    grande = (
        limite_linhas is not None
        and (os.path.getsize(caminho) > TAMANHO_LEITURA_EM_BLOCOS or aviso_memoria is not None)
    )
    if grande and limite_linhas is not None:
        if aviso_memoria:
            avisos.append({
                "tipo": "Amostragem por limite de memória",
                "severidade": "🟡 MÉDIA",
                "mensagem": aviso_memoria,
            })
        df, total_arquivo = _ler_csv_amostrado(caminho, encoding, sep, inicio, limite_linhas)
        df.attrs["linhas_originais"] = total_arquivo
        if aviso_memoria:
            df.attrs["motivo_amostragem"] = aviso_memoria
        _avisar_se_encoding_teve_substituicao(df, avisos, encoding)
        df = layout_mod.converter_datas_iso(df)
        if detectar:
            df, lay = _preparar_corpo(df, avisos)
            lay.linha_cabecalho, lay.separador, lay.encoding = inicio, sep, encoding
            _anexar_layout(df, lay)
            df.attrs["linhas_originais"] = total_arquivo
        return df

    df = (
        _ler_csv(caminho, encoding, sep) if inicio == 0
        # O pyarrow não aceita `skiprows`: arquivo com preâmbulo cai no engine C.
        else pd.read_csv(
            caminho, encoding=encoding, sep=sep, skiprows=inicio, low_memory=False,
            encoding_errors="replace",
        )
    )
    _avisar_se_encoding_teve_substituicao(df, avisos, encoding)
    # Aplicado aos dois caminhos de propósito. Os engines discordam sobre data
    # ISO — o pyarrow converte o que tem hora, o C não converte nada — e o
    # mesmo arquivo saía com tipos diferentes só por ter, ou não, um título em
    # cima. Aqui os dois terminam no mesmo lugar.
    df = layout_mod.converter_datas_iso(df)
    if detectar:
        df, lay = _preparar_corpo(df, avisos)
        lay.linha_cabecalho = inicio
        lay.separador = sep
        lay.encoding = encoding
        _anexar_layout(df, lay)
    logger.info(f"CSV carregado com separador {sep!r} | Shape: {df.shape}")
    return df


def _carregar_aba_com_layout(
    caminho: str, aba: str, engine: ExcelEngine, detectar: bool,
    linha_cabecalho: int | None, limite_linhas: int | None = None,
) -> pd.DataFrame:
    inicio = linha_cabecalho or 0
    avisos: list = []
    if detectar and linha_cabecalho is None:
        bruto = pd.read_excel(
            caminho, sheet_name=aba, engine=engine, header=None,
            nrows=_LINHAS_INSPECAO_LAYOUT,
        )
        if isinstance(bruto, pd.DataFrame):
            inicio, avisos = layout_mod.detectar_linha_cabecalho(bruto)

    aviso_memoria = _amostragem_por_memoria(caminho, "excel") if limite_linhas else None
    if aviso_memoria:
        avisos.append({
            "tipo": "Amostragem por limite de memória",
            "severidade": "🟡 MÉDIA",
            "mensagem": aviso_memoria,
        })
    if aviso_memoria and engine == "openpyxl":
        df, total_arquivo = _ler_xlsx_amostrado(caminho, aba, inicio, limite_linhas or 1)
        df.attrs["linhas_originais"] = total_arquivo
        df.attrs["motivo_amostragem"] = aviso_memoria
    else:
        lido = pd.read_excel(
            caminho, sheet_name=aba, engine=engine, header=inicio,
            nrows=limite_linhas if aviso_memoria else None,
        )
        df = lido if isinstance(lido, pd.DataFrame) else pd.DataFrame()
    if aviso_memoria and engine != "openpyxl":
        # Motores Excel não oferecem contagem barata e uniforme para todos os
        # formatos. A saída deixa explícito que são as primeiras linhas e que
        # o total não foi contabilizado.
        df.attrs["motivo_amostragem"] = (
            aviso_memoria + " Neste formato legado, foram lidas as primeiras linhas; "
            "a amostra não é uniforme e o total não foi contabilizado."
        )
        df.attrs["linhas_originais_desconhecidas"] = True
    if detectar and not df.empty:
        df, lay = _preparar_corpo(df, avisos)
        lay.linha_cabecalho = inicio
        _anexar_layout(df, lay)
    return df


def _nomes_colunas_excel(cabecalho: tuple[object, ...]) -> list[str]:
    """Reproduz nomes legíveis e únicos sem depender do leitor completo."""
    usados: dict[str, int] = {}
    nomes: list[str] = []
    for indice, valor in enumerate(cabecalho):
        base = f"Unnamed: {indice}" if valor is None else str(valor)
        repeticao = usados.get(base, 0)
        usados[base] = repeticao + 1
        nomes.append(base if repeticao == 0 else f"{base}.{repeticao}")
    return nomes


def _ler_xlsx_amostrado(
    caminho: str, aba: str, inicio: int, limite: int
) -> tuple[pd.DataFrame, int]:
    """Conta e amostra uma aba XLSX inteira sem materializá-la em memória.

    A reserva tem probabilidade uniforme por linha, como no caminho de CSV.
    Diferentemente de ``nrows``, não privilegia os primeiros dias/meses de um
    export ordenado.
    """
    from openpyxl import load_workbook

    livro = load_workbook(caminho, read_only=True, data_only=True)
    try:
        planilha = livro[aba]
        linhas = planilha.iter_rows(min_row=inicio + 1, values_only=True)
        cabecalho = next(linhas, None)
        if cabecalho is None:
            return pd.DataFrame(), 0
        reserva: list[tuple[object, ...]] = []
        gerador = np.random.default_rng(42)
        total = 0
        for linha in linhas:
            total += 1
            if len(reserva) < limite:
                reserva.append(linha)
                continue
            indice = int(gerador.integers(0, total))
            if indice < limite:
                reserva[indice] = linha
        logger.info(
            f"Leitura XLSX em streaming: {total:,} linhas na aba, "
            f"{len(reserva):,} sorteadas para análise."
        )
        return pd.DataFrame(reserva, columns=_nomes_colunas_excel(cabecalho)), total
    finally:
        livro.close()


def _carregar_parquet(caminho: str) -> pd.DataFrame:
    """Parquet já vem tipado e sem preâmbulo — não há layout a detectar.

    Entrou na lista de entrada porque a ferramenta já exportava Parquet e não
    lia: quem guardou a camada Silver em Parquet não conseguia perfilá-la.
    """
    try:
        return pd.read_parquet(caminho)
    except Exception as e:
        raise FileFormatError(f"Falha ao ler o Parquet '{caminho}': {e}") from e


def carregar_arquivo(
    caminho: str,
    aba_excel: str | int | None = 0,
    detectar_layout: bool = True,
    linha_cabecalho: int | None = None,
    limite_linhas: int | None = None,
) -> tuple[pd.DataFrame, str]:
    """Carrega um arquivo, corrigindo o layout de planilha feita à mão.

    `detectar_layout=False` volta ao comportamento cru (cabeçalho na primeira
    linha, nada removido); `linha_cabecalho` força a linha manualmente.
    """
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: '{caminho}'")

    extensao, compactacao = _partes_da_extensao(caminho)
    if compactacao == ".zip":
        _validar_zip(caminho)
    nome_base = os.path.basename(caminho)
    for sufixo in (compactacao, extensao):
        if sufixo and nome_base.lower().endswith(sufixo):
            nome_base = nome_base[: -len(sufixo)]

    formato = formato_de(caminho)
    if formato == "parquet":
        df = _carregar_parquet(caminho)
        logger.info(f"Parquet carregado | Shape: {df.shape}")
        return df, nome_base

    if formato == "texto":
        try:
            df = _carregar_csv_com_layout(
                caminho, detectar_layout, linha_cabecalho, limite_linhas
            )
        except FileFormatError:
            raise
        except Exception as e:
            raise FileFormatError(f"Falha ao ler o CSV '{caminho}': {e}") from e
        return df, nome_base

    if formato != "excel":
        raise FileFormatError(
            f"Extensão '{extensao or 'sem extensão'}' não suportada. Use: "
            f"{', '.join(EXTENSOES_SUPORTADAS)}."
        )

    engine = _ENGINES_EXCEL[extensao]
    try:
        xl = pd.ExcelFile(caminho, engine=engine)
        abas = xl.sheet_names
        if isinstance(aba_excel, int):
            if not -len(abas) <= aba_excel < len(abas):
                raise FileFormatError(
                    f"Aba de índice {aba_excel} não existe em '{caminho}' "
                    f"({len(abas)} aba(s): {', '.join(map(str, abas))})."
                )
            aba_alvo = abas[aba_excel]
        else:
            aba_alvo = str(aba_excel)
        df = _carregar_aba_com_layout(
            caminho, str(aba_alvo), engine, detectar_layout, linha_cabecalho, limite_linhas
        )
        nome_tabela = f"{nome_base}__{aba_alvo}"
        logger.info(f"[{extensao}] Aba '{aba_alvo}' carregada | Shape: {df.shape}")
        return df, nome_tabela
    except FileFormatError:
        raise
    except Exception as e:
        raise FileFormatError(f"Falha ao ler '{caminho}' ({extensao}): {e}") from e


def listar_abas(caminho: str) -> list[str]:
    """Nomes das abas de um Excel — vazio para qualquer outro formato.

    Existe para a CLI poder avisar que há abas fora da análise: perfilar
    silenciosamente só a primeira aba de um arquivo com cinco é a forma mais
    fácil de alguém concluir coisa errada sobre os dados.
    """
    extensao, _ = _partes_da_extensao(caminho)
    if extensao not in _ENGINES_EXCEL or not os.path.exists(caminho):
        return []
    try:
        return [str(aba) for aba in pd.ExcelFile(caminho, engine=_ENGINES_EXCEL[extensao]).sheet_names]
    except Exception:
        return []


def carregar_todas_abas_excel(
    caminho: str, detectar_layout: bool = True, limite_linhas: int | None = None,
) -> list[tuple[pd.DataFrame, str]]:
    if not os.path.exists(caminho):
        raise FileNotFoundError(f"Arquivo não encontrado: '{caminho}'")

    extensao, _ = _partes_da_extensao(caminho)
    engine = _ENGINES_EXCEL.get(extensao, "openpyxl")
    nome_base = os.path.splitext(os.path.basename(caminho))[0]

    try:
        xl = pd.ExcelFile(caminho, engine=engine)
        resultado = []
        for aba in xl.sheet_names:
            df_aba = _carregar_aba_com_layout(
                caminho, str(aba), engine, detectar_layout, None, limite_linhas
            )
            logger.info(f"Aba '{aba}' carregada | Shape: {df_aba.shape}")
            resultado.append((df_aba, f"{nome_base}__{aba}"))
        return resultado
    except Exception as e:
        raise FileFormatError(f"Falha ao ler '{caminho}' ({extensao}): {e}") from e
